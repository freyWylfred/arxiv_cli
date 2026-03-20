# ========================================
# arxiv_cli.py
# ========================================
# 概要:
#   arXiv から論文を検索・取得し、PDFをダウンロードして
#   Azure OpenAI API を使用して日本語要約を生成するスクリプト
#
# 主な機能:
#   - arXiv API を使用した論文検索
#   - 指定日付範囲の論文をフィルタリング
#   - PDF ファイルの自動ダウンロード
#   - Azure OpenAI API による日本語要約生成（オプション）
#   - Excel ファイルへの結果出力（ハイパーリンク付き）
#   - 処理ログの出力
#
# 使用方法:
#   1. config.ini を設定
#   2. python arxiv_cli.py を実行
#
# 必要なライブラリ:
#   pip install openai==0.28.1 feedparser PyMuPDF requests pandas openpyxl
#
# ライセンス: MIT License
# ========================================

# ========================================
# ライブラリのインポート
# ========================================
# 標準ライブラリ
import sys               # exe判定・パス解決用
import urllib.parse      # URLエンコーディング用
from urllib.parse import urlparse  # URLのパース（分解）用
import os                # ファイルパス操作用
import time              # リトライ用待機時間
import logging           # ログ出力用
import configparser      # INIファイル読み込み用
from datetime import datetime, timedelta  # 日付フィルタリング用

# サードパーティライブラリ
import feedparser        # arXiv APIからのRSSフィード解析用
import requests          # PDFダウンロード用
import fitz              # PyMuPDF - PDFからテキスト抽出用
import openai            # OpenAI APIでの要約生成用
import pandas as pd      # Excel出力用
from openpyxl import load_workbook  # 既存Excelファイル操作用
from openpyxl.worksheet.hyperlink import Hyperlink  # ハイパーリンク用

# ========================================
# exe実行時のパス解決
# ========================================
def get_app_dir():
    """
    アプリケーションの実行ディレクトリを取得する関数
    PyInstallerでexe化された場合はexeの配置場所を返す
    """
    if getattr(sys, 'frozen', False):
        # PyInstallerでビルドされたexeの場合
        return os.path.dirname(sys.executable)
    else:
        # 通常のPythonスクリプトの場合
        return os.path.dirname(os.path.abspath(__file__))

# カレントディレクトリをアプリケーションディレクトリに設定
# （exe実行時にconfig.ini等の相対パスが正しく解決されるようにする）
os.chdir(get_app_dir())

# ========================================
# ログ設定
# ========================================
# ログの出力設定
# - ファイル出力: arxiv_process.log（UTF-8エンコーディング）
# - コンソール出力: 標準出力
# - ログレベル: INFO以上を記録
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('arxiv_process.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# ========================================
# デフォルト設定ファイル生成
# ========================================
DEFAULT_CONFIG_CONTENT = """\
[OpenAI]
# OpenAI要約機能を有効にする場合は true に変更
use_openai = false

# Azure OpenAI APIタイプ
api_type = azure

# Azure OpenAI エンドポイント
endpoint = https://your-resource-name.openai.azure.com/

# Azure OpenAI APIバージョン
api_version = 2024-02-15-preview

# Azure OpenAI APIキー
api_key =

# Azure OpenAI デプロイメント名
deployment_name = gpt-4

[DateRange]
# 今日の日付のみを処理するかどうか (true/false)
# true の場合 start_date と end_date の設定は無視されます
today_only = true

# 検索開始日 (YYYY/MM/DD形式)
start_date = 2026/01/01

# 検索終了日 (YYYY/MM/DD形式)
end_date = 2026/01/01

[Search]
# 検索クエリ (arXiv API形式)
query = all:"model extraction" OR all:"prompt injection"

# 最大取得件数
max_results = 100

[Files]
# Excel出力ファイル名
excel_file = arxiv_summaries.xlsx
"""


def create_default_config(config_file='config.ini'):
    """
    デフォルトの config.ini を自動生成する関数

    Args:
        config_file (str): 生成する設定ファイルのパス
    """
    try:
        with open(config_file, 'w', encoding='utf-8') as f:
            f.write(DEFAULT_CONFIG_CONTENT)
        logging.info(f"デフォルト設定ファイルを自動生成しました: {config_file}")
        print(f"📝 設定ファイルが見つからなかったため、デフォルトで自動生成しました: {config_file}")
    except Exception as e:
        logging.error(f"設定ファイルの自動生成に失敗しました: {str(e)}")
        print(f"❌ 設定ファイルの自動生成に失敗しました: {str(e)}")
        exit(1)


# ========================================
# 設定ファイル読み込み
# ========================================
def load_config(config_file='config.ini'):
    """
    INIファイルから設定を読み込む関数
    ファイルが存在しない場合はデフォルト設定で自動生成する

    Args:
        config_file (str): 設定ファイルのパス

    Returns:
        dict: 設定情報を格納した辞書
    """
    if not os.path.exists(config_file):
        create_default_config(config_file)

    config = configparser.ConfigParser()

    # INIファイルを読み込み（UTF-8エンコーディング）
    try:
        config.read(config_file, encoding='utf-8')
    except Exception as e:
        logging.error(f"INIファイルの読み込みに失敗しました: {str(e)}")
        print(f"❌ INIファイルの読み込みに失敗しました: {str(e)}")
        exit(1)

    # セクションの存在確認
    required_sections = ['DateRange', 'Search', 'Files']
    for section in required_sections:
        if not config.has_section(section):
            logging.error(f"必須セクション [{section}] がINIファイルに見つかりません")
            print(f"❌ エラー: 設定ファイルに [{section}] セクションがありません")
            exit(1)

    try:
        # 「今日のだけ」オプションの取得
        today_only = config.getboolean('DateRange', 'today_only', fallback=False)

        # 日付範囲の取得
        if today_only:
            # 「今日のだけ」が有効な場合、今日の日付を使用
            start_date = datetime.now().date()
            end_date = datetime.now().date()
            start_date_str = start_date.strftime("%Y/%m/%d")
            end_date_str = end_date.strftime("%Y/%m/%d")
            logging.info("「今日のだけ」モードが有効です。今日の日付を使用します。")
        else:
            # 通常モード：INIファイルから日付を読み込み
            start_date_str = config.get('DateRange', 'start_date').strip()
            end_date_str = config.get('DateRange', 'end_date').strip()
            
            # 日付のバリデーション
            start_date = datetime.strptime(start_date_str, "%Y/%m/%d").date()
            end_date = datetime.strptime(end_date_str, "%Y/%m/%d").date()

            if start_date > end_date:
                logging.error("start_dateがend_dateより後の日付になっています")
                print("❌ エラー: start_dateはend_date以前の日付を指定してください")
                exit(1)

        # 検索設定の取得
        query = config.get('Search', 'query').strip()
        if not query:
            logging.error("検索クエリが空です")
            print("❌ エラー: [Search] query が空です。検索キーワードを設定してください")
            exit(1)

        max_results = config.getint('Search', 'max_results')
        if max_results <= 0:
            logging.error(f"max_results が不正な値です: {max_results}")
            print("❌ エラー: [Search] max_results は1以上の整数を指定してください")
            exit(1)

        # ファイル設定の取得
        excel_file = config.get('Files', 'excel_file').strip()

        # OpenAI設定の取得（デフォルト: 無効）
        use_openai = config.getboolean('OpenAI', 'use_openai', fallback=False)

        # OpenAI詳細設定の取得（config.ini優先、なければ環境変数、なければデフォルト値）
        openai_api_type = config.get('OpenAI', 'api_type', fallback=os.environ.get("OPENAI_API_TYPE", "azure"))
        openai_endpoint = config.get('OpenAI', 'endpoint', fallback=os.environ.get("AZURE_OPENAI_ENDPOINT", "https://your-resource-name.openai.azure.com/"))
        openai_api_version = config.get('OpenAI', 'api_version', fallback=os.environ.get("AZURE_OPENAI_API_VERSION", "2024-02-15-preview"))
        openai_api_key = config.get('OpenAI', 'api_key', fallback=os.environ.get("AZURE_OPENAI_API_KEY", ""))
        openai_deployment_name = config.get('OpenAI', 'deployment_name', fallback="gpt-4")

        # 設定値の検証ログ
        logging.info("="*60)
        logging.info("INIファイルから読み込んだ設定:")
        logging.info(f"  検索期間: {start_date_str} ～ {end_date_str}")
        logging.info(f"  検索クエリ: {query}")
        logging.info(f"  最大取得件数: {max_results}")
        logging.info(f"  Excel出力ファイル: {excel_file}")
        logging.info(f"  OpenAI要約: {'有効' if use_openai else '無効'}")
        if use_openai:
            logging.info(f"  OpenAI APIタイプ: {openai_api_type}")
            logging.info(f"  OpenAI エンドポイント: {openai_endpoint}")
            logging.info(f"  OpenAI デプロイメント名: {openai_deployment_name}")
        logging.info("="*60)

        return {
            'start_date': start_date,
            'end_date': end_date,
            'query': query,
            'max_results': max_results,
            'excel_file': excel_file,
            'use_openai': use_openai,
            'openai_api_type': openai_api_type,
            'openai_endpoint': openai_endpoint,
            'openai_api_version': openai_api_version,
            'openai_api_key': openai_api_key,
            'openai_deployment_name': openai_deployment_name
        }

    except Exception as e:
        logging.error(f"設定ファイルの読み込みエラー: {str(e)}")
        print(f"❌ 設定ファイルの読み込みエラー: {str(e)}")
        print("config.iniファイルの形式を確認してください。")
        exit(1)

# ========================================
# OpenAI API設定関数
# ========================================
def setup_openai(config):
    """
    config辞書からOpenAI APIを設定する関数

    Args:
        config (dict): load_config()から返された設定辞書
    """
    openai.api_type = config.get('openai_api_type', 'azure')
    openai.api_base = config.get('openai_endpoint', '')
    openai.api_version = config.get('openai_api_version', '2024-02-15-preview')
    openai.api_key = config.get('openai_api_key', '')

    # API キーの存在確認（要約機能を使用する場合のみ必要）
    if config.get('use_openai', False) and not openai.api_key:
        logging.warning("Azure OpenAI API キーが設定されていません。")
        logging.warning("config.ini の api_key または環境変数 AZURE_OPENAI_API_KEY を設定してください。")

# ========================================
# PDF要約関数
# ========================================
def summarize_pdf(pdf_path, max_pages=5, use_openai=False, deployment_name="gpt-4"):
    """
    PDFファイルからテキストを抽出し、OpenAI APIを使って日本語で要約する関数

    Args:
        pdf_path (str): 要約したいPDFファイルのパス
        max_pages (int): 要約対象の最大ページ数（トークン制限対策）
        use_openai (bool): OpenAI APIを使用するかどうか（デフォルト: False）
        deployment_name (str): Azure OpenAIのデプロイメント名（デフォルト: gpt-4）

    Returns:
        tuple: (要約テキスト, 使用トークン数の辞書)
               使用トークン数の辞書は {'prompt_tokens': int, 'completion_tokens': int, 'total_tokens': int}
    """
    try:
        # PyMuPDFを使ってPDFファイルを開く
        doc = fitz.open(pdf_path)

        # 全ページのテキストを結合するための変数
        full_text = ""

        # PDFの最初のmax_pagesページからテキストを抽出（トークン制限対策）
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            full_text += page.get_text()

        doc.close()

        # テキストが空の場合
        if not full_text.strip():
            return "テキストを抽出できませんでした。", {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0}

        # テキストを最大8000文字に制限（トークン制限対策）
        full_text = full_text[:8000]

        # ========================================
        # OpenAI API呼び出し
        # ========================================
        if use_openai:
            # Azure OpenAI APIを使って要約を生成
            response = openai.ChatCompletion.create(
                engine=deployment_name,  # config.iniで設定されたデプロイメント名
                messages=[
                    {"role": "user", "content": "次の論文の冒頭部分を日本語で簡潔に要約してください（200文字程度）：\n\n" + full_text}
                ],
                max_tokens=500,
                temperature=0.5
            )

            # APIレスポンスから要約テキストと使用トークン数を取得
            summary = response.choices[0].message['content']
            usage = response.usage
            token_usage = {
                'prompt_tokens': usage.prompt_tokens,
                'completion_tokens': usage.completion_tokens,
                'total_tokens': usage.total_tokens
            }
            return summary, token_usage
        else:
            # OpenAI API未使用時の暫定的な要約
            summary = "【要約未実施】OpenAI APIが無効に設定されています。config.iniのuse_openai=trueで有効化できます。"
            # API未使用の場合は概算トークン数を返す（文字数の1/4程度）
            estimated_tokens = len(full_text) // 4
            token_usage = {
                'prompt_tokens': estimated_tokens,
                'completion_tokens': 100,  # 要約の概算トークン数
                'total_tokens': estimated_tokens + 100
            }
            return summary, token_usage

    except Exception as e:
        return f"要約生成エラー: {str(e)}", {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0}

# ========================================
# Excel保存関数
# ========================================
def add_hyperlinks_to_sheet(workbook, sheet_name, url_column='PDF URL'):
    """
    指定したシートのURLカラムにハイパーリンクを設定する関数

    Args:
        workbook: openpyxlのワークブックオブジェクト
        sheet_name (str): シート名
        url_column (str): URLが含まれるカラム名
    """
    sheet = workbook[sheet_name]

    # ヘッダー行からURLカラムのインデックスを取得
    url_col_index = None
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value == url_column:
            url_col_index = col_idx
            break

    if url_col_index is None:
        logging.warning(f"URLカラム '{url_column}' が見つかりません")
        return

    # 2行目以降（データ行）にハイパーリンクを設定
    for row_idx in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=url_col_index)
        url_value = cell.value
        if url_value and url_value.startswith('http'):
            cell.hyperlink = url_value
            cell.style = 'Hyperlink'  # ハイパーリンクスタイルを適用

    logging.info(f"シート '{sheet_name}' のURLにハイパーリンクを設定しました")


def save_to_excel(data_list, sheet_name, excel_file="arxiv_summaries.xlsx", max_retries=3):
    """
    要約結果をExcelファイルに保存する関数

    Args:
        data_list (list): 保存するデータのリスト（辞書形式）
        sheet_name (str): シート名（日付）
        excel_file (str): Excelファイル名
        max_retries (int): 最大リトライ回数
    """
    if not data_list:
        logging.warning("保存するデータがありません。")
        return False

    # シート名の長さチェック（Excelの制限は31文字）
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
        logging.warning(f"シート名が長すぎるため切り詰めました: {sheet_name}")

    # シート名に使用できない文字を置換
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '-')

    for attempt in range(max_retries):
        try:
            # データをDataFrameに変換
            df = pd.DataFrame(data_list)

            # Excelファイルが存在するか確認
            if os.path.exists(excel_file):
                try:
                    # 既存のExcelファイルを読み込み
                    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # 既存のワークブックを取得
                        workbook = writer.book

                        # シートが既に存在するか確認
                        if sheet_name in workbook.sheetnames:
                            # 既存シートのデータを読み込み
                            existing_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                            # 新しいデータを追加
                            combined_df = pd.concat([existing_df, df], ignore_index=True)
                            # シートを削除して再作成
                            del workbook[sheet_name]
                            combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            # シートを先頭に移動
                            workbook.move_sheet(sheet_name, offset=-len(workbook.sheetnames)+1)
                            # URLにハイパーリンクを設定
                            add_hyperlinks_to_sheet(workbook, sheet_name)
                            logging.info(f"既存シート '{sheet_name}' にデータを追加しました（先頭に移動）")
                        else:
                            # 新しいシートとして追加
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            # シートを先頭に移動
                            workbook.move_sheet(sheet_name, offset=-len(workbook.sheetnames)+1)
                            # URLにハイパーリンクを設定
                            add_hyperlinks_to_sheet(workbook, sheet_name)
                            logging.info(f"新規シート '{sheet_name}' を作成しました（先頭に配置）")
                except PermissionError:
                    if attempt < max_retries - 1:
                        logging.warning(f"Excelファイルが開かれています。3秒後にリトライします... ({attempt + 1}/{max_retries})")
                        time.sleep(3)
                        continue
                    else:
                        error_msg = f"❌ Excelファイルが開かれています。ファイルを閉じてから再実行してください: {excel_file}"
                        logging.error(error_msg)
                        print(f"\n{error_msg}")
                        return False
            else:
                # 新規 Excelファイルを作成
                df.to_excel(excel_file, sheet_name=sheet_name, index=False, engine='openpyxl')
                logging.info(f"新規Excelファイル '{excel_file}' を作成しました")

                # ハイパーリンクを設定するためにファイルを再度開く
                wb = load_workbook(excel_file)
                add_hyperlinks_to_sheet(wb, sheet_name)
                wb.save(excel_file)
                wb.close()

            print(f"\n✅ Excelに保存しました: {excel_file} (シート: {sheet_name})")
            logging.info(f"Excel保存成功: {len(data_list)}件のデータ")
            return True

        except MemoryError:
            error_msg = "❌ メモリ不足エラー。データ量が多すぎます。"
            logging.error(error_msg)
            print(f"\n{error_msg}")
            return False

        except Exception as e:
            if attempt < max_retries - 1:
                logging.warning(f"Excel保存エラー。リトライします... ({attempt + 1}/{max_retries}): {str(e)}")
                time.sleep(2)
            else:
                error_msg = f"❌ Excel保存エラー: {str(e)}"
                logging.error(error_msg)
                print(f"\n{error_msg}")
                return False

    return False

# ========================================
# メイン処理関数
# ========================================
def process_date(target_datetime, feed, excel_file, use_openai=False, deployment_name="gpt-4"):
    """
    指定された日付の論文を処理する関数

    Args:
        target_datetime (date): 処理対象の日付
        feed: arXiv APIから取得したフィード
        excel_file (str): Excel出力ファイル名
        use_openai (bool): OpenAI APIを使用するかどうか
        deployment_name (str): Azure OpenAIのデプロイメント名

    Returns:
        tuple: (処理した論文数, トークン使用量の辞書)
    """
    target_date = target_datetime.strftime("%Y/%m/%d")
    sheet_name = target_date.replace("/", "-")

    # トークン使用量の累計
    total_token_usage = {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0}

    logging.info(f"処理開始: {target_date} (Excel出力先: {excel_file}, シート名: {sheet_name})")
    print(f"\n{'='*60}")
    print(f"検索対象日付: {target_date}")
    print(f"Excel出力先: {excel_file}")
    print(f"{'='*60}")

    # ========================================
    # 重複チェック：既に調査済みかどうかを確認
    # ========================================
    if os.path.exists(excel_file):
        try:
            # Excelファイルを開いてシート名をチェック
            workbook = load_workbook(excel_file, read_only=True)

            if sheet_name in workbook.sheetnames:
                # 既にシートが存在する場合
                workbook.close()
                print(f"⚠️  既に調査済みです！日付: {target_date} (シート: {sheet_name})")
                logging.info(f"既に調査済み: {target_date} (シート: {sheet_name})")
                print(f"この日付をスキップします。")
                return 0, total_token_usage

            workbook.close()
            logging.info(f"新規調査対象: {target_date}")

        except Exception as e:
            logging.warning(f"Excelファイルチェック中にエラー: {str(e)}")
            print(f"⚠️  Excelファイルチェック中にエラーが発生しましたが、処理を続行します。")
    else:
        logging.info(f"Excelファイル未作成。新規調査: {target_date}")
        print(f"新規調査を開始します: {target_date}")

    print(f"\n論文の取得と要約を開始します...")

    # ========================================
    # PDFダウンロード用フォルダの準備
    # ========================================
    # 日付フォルダ名のベース（YYYYMMDD形式）
    pdf_folder_base = target_date.replace("/", "")  # "20260201" 形式
    pdf_folder = pdf_folder_base  # 初期値

    # ========================================
    # 各論文エントリの処理ループ
    # ========================================
    found_count = 0
    results_data = []  # Excel出力用のデータリスト

    for entry in feed.entries:
        try:
            # ----------------------------------------
            # 日付フィルタリング
            # ----------------------------------------
            # 論文の投稿日時を取得してdate型に変換
            published = datetime.strptime(entry.published, "%Y-%m-%dT%H:%M:%SZ").date()

            # 指定した日付と一致しない場合は、この論文をスキップして次へ
            if published != target_datetime:
                continue

            found_count += 1
            print(f"\n{'='*60}")
            print(f"論文 #{found_count} を処理中...")

            # ----------------------------------------
            # 最初の論文が見つかった時点でフォルダを作成
            # ----------------------------------------
            if found_count == 1:
                # 論文が見つかったので通常のフォルダ名
                if not os.path.exists(pdf_folder):
                    os.makedirs(pdf_folder)
                    logging.info(f"PDFダウンロード用フォルダを作成: {pdf_folder}")
                    print(f"PDFダウンロード用フォルダを作成しました: {pdf_folder}")
                else:
                    logging.info(f"既存のフォルダを使用: {pdf_folder}")

            # ----------------------------------------
            # arXiv IDとPDF URLの取得
            # ----------------------------------------
            arxiv_id = entry.id.split('/abs/')[-1]
            pdf_url = f'https://arxiv.org/pdf/{arxiv_id}.pdf'

            # ----------------------------------------
            # ファイル名の処理
            # ----------------------------------------
            parsed = urlparse(pdf_url)
            filename = os.path.basename(parsed.path)

            # バージョン番号を除いたファイル名を生成
            # arXiv のファイル名は "2506.12345v2.pdf" 形式
            # rsplit で最後の 'v' のみを対象にし、ID 自体に 'v' を含む場合の誤動作を防ぐ
            name_no_ext = filename.rsplit('.', 1)[0]  # "2506.12345v2"
            if 'v' in name_no_ext and name_no_ext.rsplit('v', 1)[-1].isdigit():
                base_id = name_no_ext.rsplit('v', 1)[0]
                filename_no_version = base_id + '.pdf'
            else:
                filename_no_version = filename

            # PDFファイルの保存パス（日付フォルダ内に保存）
            pdf_file_path = os.path.join(pdf_folder, filename)
            pdf_file_path_no_version = os.path.join(pdf_folder, filename_no_version)

            # ----------------------------------------
            # PDFのダウンロード
            # ----------------------------------------
            print(f"PDFをダウンロード中: {pdf_url}")
            logging.info(f"PDFダウンロード開始: {arxiv_id}")

            # PDFダウンロード（リトライ付き）
            download_success = False
            for download_attempt in range(3):
                try:
                    response = requests.get(pdf_url, timeout=30)
                    response.raise_for_status()

                    # Content-Typeの確認
                    content_type = response.headers.get('Content-Type', '')
                    if 'pdf' not in content_type.lower() and 'octet-stream' not in content_type.lower():
                        logging.warning(f"PDFではない可能性があります: {content_type}")

                    # ファイルサイズチェック（最小サイズ）
                    if len(response.content) < 1000:
                        logging.warning(f"ダウンロードしたファイルが小さすぎます: {len(response.content)}バイト")
                        if download_attempt < 2:
                            time.sleep(2)
                            continue

                    # 日付フォルダ内にPDFを保存
                    with open(pdf_file_path, 'wb') as f:
                        f.write(response.content)

                    download_success = True
                    file_size_kb = len(response.content) / 1024
                    logging.info(f"PDFダウンロード成功: {file_size_kb:.2f}KB -> {pdf_file_path}")
                    print(f"保存完了: {pdf_file_path} ({file_size_kb:.2f}KB)")
                    break

                except requests.exceptions.Timeout:
                    logging.warning(f"タイムアウト ({download_attempt + 1}/3)")
                    if download_attempt < 2:
                        time.sleep(3)
                except requests.exceptions.RequestException as e:
                    logging.error(f"ダウンロードエラー ({download_attempt + 1}/3): {str(e)}")
                    if download_attempt < 2:
                        time.sleep(3)

            if not download_success:
                logging.error(f"PDFダウンロード失敗: {arxiv_id}")
                print(f"❌ PDFダウンロード失敗: {arxiv_id}")
                continue

            # ----------------------------------------
            # PDF要約の生成
            # ----------------------------------------
            print("要約を生成中...")
            logging.info(f"要約生成開始: {arxiv_id}")
            summary, token_usage = summarize_pdf(pdf_file_path, use_openai=use_openai, deployment_name=deployment_name)  # タプルで受け取る

            # トークン使用量を累計
            total_token_usage['prompt_tokens'] += token_usage['prompt_tokens']
            total_token_usage['completion_tokens'] += token_usage['completion_tokens']
            total_token_usage['total_tokens'] += token_usage['total_tokens']

            # 要約が空またはエラーの場合
            if not summary or "エラー" in summary:
                logging.warning(f"要約生成に問題が発生しました: {arxiv_id}")
            else:
                logging.info(f"要約生成完了: {len(summary)}文字, トークン使用: {token_usage['total_tokens']}")

            # ----------------------------------------
            # 結果の出力
            # ----------------------------------------
            print('\nTitle:', entry.title)
            print('arXiv ID:', arxiv_id)
            print('PDF URL:', pdf_url)
            print('PDF filename:', pdf_file_path)
            print('PDF filename (no version):', pdf_file_path_no_version)
            print('\n【要約】')
            print(summary)
            print('='*60)

            # ----------------------------------------
            # Excel保存用データに追加
            # ----------------------------------------
            result_data = {
                '投稿日': published.strftime("%Y-%m-%d"),
                'タイトル': entry.title,
                'arXiv ID': arxiv_id,
                'PDF URL': pdf_url,
                'ファイル名': filename_no_version,
                '要約': summary
            }
            results_data.append(result_data)

        except Exception as e:
            logging.error(f"論文エントリ処理中にエラー: {str(e)}")
            print(f"エラーが発生しました: {str(e)}")
            continue

    # ========================================
    # Excel保存処理
    # ========================================
    if found_count == 0:
        # ========================================
        # 該当論文なし - フォルダ名に「(該当論文無し)」を追加
        # ========================================
        pdf_folder = pdf_folder_base + "(該当論文無し)"

        # フォルダを作成（該当論文無しの記録用）
        if not os.path.exists(pdf_folder):
            os.makedirs(pdf_folder)
            logging.info(f"該当論文無しフォルダを作成: {pdf_folder}")
            print(f"\n該当論文無しフォルダを作成しました: {pdf_folder}")

        message = f"\n{target_date} に投稿された論文は見つかりませんでした。"
        logging.info(message)
        print(message)

        # 該当論文がない場合もExcelにシートを追加して記録
        no_data_message = {
            '投稿日': target_date,
            'タイトル': 'この日は該当論文がありませんでした',
            'arXiv ID': '-',
            'PDF URL': '-',
            'ファイル名': '-',
            '要約': f'{target_date} には検索条件に一致する論文が投稿されていません。'
        }
        results_data = [no_data_message]

        excel_success = save_to_excel(results_data, sheet_name, excel_file)

        if excel_success:
            print(f"✅ 「該当なし」の記録をExcelに保存しました。")
        else:
            print("⚠️  Excel保存に失敗しました。")

    else:
        print(f"\n処理完了: {found_count}件の論文を処理しました。")
        logging.info(f"処理完了: {found_count}件")

        # Excelに保存（シート名は日付）
        if results_data:
            excel_success = save_to_excel(results_data, sheet_name, excel_file)

            if not excel_success:
                # Excel保存失敗時、CSVにバックアップ
                try:
                    csv_filename = f"arxiv_backup_{sheet_name}.csv"
                    df_backup = pd.DataFrame(results_data)
                    df_backup.to_csv(csv_filename, index=False, encoding='utf-8-sig')
                    logging.info(f"CSVバックアップ作成: {csv_filename}")
                    print(f"⚠️  Excel保存に失敗したため、CSVに保存しました: {csv_filename}")
                except Exception as e:
                    logging.error(f"CSVバックアップ失敗: {str(e)}")
                    print(f"❌ データの保存に失敗しました: {str(e)}")
        else:
            logging.warning("保存するデータがありません")

    return found_count, total_token_usage

# ========================================
# メイン実行部
# ========================================
if __name__ == "__main__":
  try:
    print("="*60)
    print("arXiv 論文取得・要約スクリプト")
    print("="*60)

    # ========================================
    # STEP 1: 設定ファイルの読み込み
    # ========================================
    print("\n設定ファイル (config.ini) を読み込んでいます...")
    config = load_config()

    # 設定値を変数に展開
    start_date = config['start_date']           # 検索開始日
    end_date = config['end_date']               # 検索終了日
    query = config['query']                     # 検索クエリ
    max_results = config['max_results']         # 最大取得件数
    excel_file = config['excel_file']           # Excel出力ファイル名
    use_openai = config['use_openai']           # OpenAI使用フラグ
    deployment_name = config['openai_deployment_name']  # デプロイメント名

    # ========================================
    # STEP 2: OpenAI APIの初期設定
    # ========================================
    if use_openai:
        setup_openai(config)

    # 設定内容をコンソールに表示
    print(f"\n✅ 設定の読み込みが完了しました")
    print(f"\n【読み込んだ設定情報】")
    print(f"  検索期間: {start_date.strftime('%Y/%m/%d')} ～ {end_date.strftime('%Y/%m/%d')}")
    print(f"  検索クエリ: {query}")
    print(f"  最大取得件数: {max_results}件")
    print(f"  出力ファイル: {excel_file}")
    print(f"  OpenAI要約: {'有効' if use_openai else '無効'}")

    # 処理対象の日数を計算
    days_count = (end_date - start_date).days + 1
    print(f"  処理対象日数: {days_count}日")

    # ログに記録
    logging.info(f"実行開始 - 検索期間: {start_date} ～ {end_date}, クエリ: {query}, 最大件数: {max_results}")

    # ========================================
    # STEP 3: arXiv API検索URLの構築
    # ========================================
    # arXiv APIのベースURL
    base = 'http://export.arxiv.org/api/query?search_query='

    # 完全なAPI URLを構築
    # - query: 検索クエリ（URLエンコード）
    # - start: 結果の開始位置（0から）
    # - max_results: 最大取得件数
    # - sortBy: ソート基準（submittedDate=投稿日）
    # - sortOrder: ソート順（descending=降順=新しい順）
    url = base + urllib.parse.quote(query) + f'&start=0&max_results={max_results}&sortBy=submittedDate&sortOrder=descending'

    logging.info(f"arXiv API URL: {url}")

    print("\narXiv APIからデータを取得中...")
    print(f"使用するクエリ: {query}")
    print(f"最大取得件数: {max_results}件")

    # ========================================
    # STEP 4: arXiv APIからフィード取得
    # ========================================
    # feedparserでRSSフィードを取得・パース
    try:
        feed = feedparser.parse(url)
    except Exception as e:
        logging.error(f"arXiv APIからのフィード取得に失敗しました: {str(e)}")
        print(f"❌ arXiv APIへの接続に失敗しました: {str(e)}")
        print("ネットワーク接続を確認してください。")
        input("\nEnterキーを押して終了...")
        exit(1)

    # フィード取得結果の検証
    if feed.bozo:
        bozo_msg = str(getattr(feed, 'bozo_exception', '不明なエラー'))
        logging.warning(f"arXiv APIフィードの解析で問題が発生しました: {bozo_msg}")
        print(f"⚠️  arXiv APIフィードの解析で問題が検出されました: {bozo_msg}")

    if not feed.entries:
        logging.warning("arXiv APIから論文が1件も取得できませんでした")
        print("⚠️  arXiv APIから論文が1件も取得できませんでした。")
        print("  - 検索クエリを確認してください")
        print("  - arXiv APIが一時的に利用できない可能性があります")

    print(f"取得した論文数: {len(feed.entries)}")

    # ========================================
    # STEP 5: 日付範囲でループ処理
    # ========================================
    # 各日付に対して論文を処理
    total_processed = 0                 # 処理した論文の総数
    current_date = start_date           # ループの現在日付

    # トークン使用量の累計（API料金計算用）
    total_token_usage = {'prompt_tokens': 0, 'completion_tokens': 0, 'total_tokens': 0}

    # 日付ごとにprocess_date()を呼び出して処理
    while current_date <= end_date:
        # 指定日付の論文を処理
        processed_count, token_usage = process_date(current_date, feed, excel_file, use_openai, deployment_name)
        total_processed += processed_count

        # トークン使用量を累計に加算
        total_token_usage['prompt_tokens'] += token_usage['prompt_tokens']
        total_token_usage['completion_tokens'] += token_usage['completion_tokens']
        total_token_usage['total_tokens'] += token_usage['total_tokens']

        # 次の日付へ進む
        current_date += timedelta(days=1)

    # ========================================
    # STEP 6: API料金計算（概算）
    # ========================================
    # GPT-4の一般的な料金（2024年時点の概算）
    # 入力トークン: $0.03 / 1,000 tokens
    # 出力トークン: $0.06 / 1,000 tokens
    input_cost = (total_token_usage['prompt_tokens'] / 1000) * 0.03
    output_cost = (total_token_usage['completion_tokens'] / 1000) * 0.06
    total_cost = input_cost + output_cost

    # 日本円換算（1ドル=150円として概算）
    total_cost_jpy = total_cost * 150

    # ========================================
    # STEP 7: 実行結果サマリーの出力
    # ========================================
    print("\n" + "="*60)
    print("全ての処理が完了しました。")
    print("="*60)
    print(f"【実行結果サマリー】")
    print(f"  処理期間: {start_date.strftime('%Y/%m/%d')} ～ {end_date.strftime('%Y/%m/%d')}")
    print(f"  処理日数: {days_count}日")
    print(f"  処理した論文総数: {total_processed}件")
    print(f"  Excel出力先: {excel_file}")
    print(f"  使用したクエリ: {query}")
    print("="*60)
    print(f"【OpenAI API 使用状況】")
    print(f"  入力トークン数: {total_token_usage['prompt_tokens']:,} tokens")
    print(f"  出力トークン数: {total_token_usage['completion_tokens']:,} tokens")
    print(f"  合計トークン数: {total_token_usage['total_tokens']:,} tokens")
    print(f"  推定料金 (USD): ${total_cost:.4f}")
    print(f"  推定料金 (JPY): ¥{total_cost_jpy:.2f} (1ドル=150円換算)")
    print("="*60)

    # ログに最終結果を記録
    logging.info(f"全処理完了: {days_count}日間で{total_processed}件の論文を処理")
    logging.info(f"Excel出力: {excel_file}, クエリ: {query}, 最大件数: {max_results}")
    logging.info(f"トークン使用: 入力{total_token_usage['prompt_tokens']}, 出力{total_token_usage['completion_tokens']}, 合計{total_token_usage['total_tokens']}")
    logging.info(f"推定料金: ${total_cost:.4f} (¥{total_cost_jpy:.2f})")
    print(f"\n✅ 結果は '{excel_file}' に保存されています。")

  except KeyboardInterrupt:
    print("\n\n処理が中断されました。")
    logging.info("ユーザーによる中断")
  except Exception as e:
    logging.error(f"予期しないエラーが発生しました: {str(e)}", exc_info=True)
    print(f"\n❌ 予期しないエラーが発生しました: {str(e)}")
    print("詳細は arxiv_process.log を確認してください。")
  finally:
    input("\nEnterキーを押して終了...")

# ========================================
# インストール手順
# ========================================
# 必要なライブラリのインストールコマンド:
#   pip install openai==0.28.1 feedparser PyMuPDF requests pandas openpyxl
#
# 注意:
#   - openai==0.28.1 は Azure OpenAI の従来APIに対応したバージョンです
#   - PyMuPDF は fitz としてインポートされます
